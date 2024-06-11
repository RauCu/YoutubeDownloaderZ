from tkinter import *
import webbrowser
from functools import partial

import os
import scrapetube
import pyperclip
import csv
import time
import re
import threading
import requests

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.styles import Font

from tkinter.scrolledtext import ScrolledText

# How to:
# https://www.python.org/ftp/python/3.8.6/python-3.8.6.exe

# pip install pyinstaller
# pip install scrapetube
# pip install pyperclip
# pip install openpyxl

# pyinstaller --onefile .\LayDanhSachVideoKenhYT.py

# from http://effbot.org/zone/tkinter-text-hyperlink.htm
class HyperlinkManager:

    def __init__(self, text):

        self.text = text

        self.text.tag_config("hyper", foreground="blue", underline=1)

        self.text.tag_bind("hyper", "<Enter>", self._enter)
        self.text.tag_bind("hyper", "<Leave>", self._leave)
        self.text.tag_bind("hyper", "<Button-1>", self._click)

        self.reset()

    def reset(self):
        self.links = {}

    def add(self, action):
        # add an action to the manager.  returns tags to use in
        # associated text widget
        tag = "hyper-%d" % len(self.links)
        self.links[tag] = action
        return "hyper", tag

    def _enter(self, event):
        self.text.config(cursor="hand2")

    def _leave(self, event):
        self.text.config(cursor="")

    def _click(self, event):
        for tag in self.text.tag_names(CURRENT):
            if tag[:6] == "hyper-":
                self.links[tag]()
                return

import subprocess
FILEBROWSER_PATH = os.path.join(os.getenv('WINDIR'), 'explorer.exe')

# Define a callback function
def callback(url):
   webbrowser.open_new_tab(url)

version = "v1.9.143 (14-Dec-2023)"

root = Tk()
root.geometry("1000x500")
root.title("Lấy danh sách video của 1 kênh Youtube - Tải hình avatar và banner - " + version)

YTChannelURL_Prefix = "https://www.youtube.com/channel/"
YTChannelID_Prefix = "UC"
resultFolder = "KetQua"

try:
    os.mkdir(resultFolder)
except FileExistsError:
    print("Folder already exists")

def explore():
    path = os.path.abspath(resultFolder)
    # explorer would choke on forward slashes
    path = os.path.normpath(path)

    if os.path.isdir(path):
        subprocess.run([FILEBROWSER_PATH, path])
    elif os.path.isfile(path):
        subprocess.run([FILEBROWSER_PATH, '/select,', path])

def exploreFile(path):
    if os.path.isdir(path):
        subprocess.run([FILEBROWSER_PATH, path])
    elif os.path.isfile(path):
        subprocess.run([FILEBROWSER_PATH, '/select,', path])        

# def getAvatarBanner():
#     webbrowser.open("https://imageyoutube.com/?")

def getVideoList(searchString):
    x = threading.Thread(target=getVideoListPrivate, args=(searchString,))
    x.start()
    # x.join()

def getVideoListPrivate(searchString):
    parts = searchString.split(YTChannelURL_Prefix, 1)
    #Output.insert(END, 'ID kênh: '+ INPUT_parts[1] +'\n')
    YTChannelID = parts[1]
    
    #channel_url = input("Nhap duong link kenh Youtube (dinh dang https://www.youtube.com/channel/xxxxx), de lay xxxxx DUNG TRANG https://http5.org/chan: ")
    # limit = input("Muon lay bao nhieu video? ")
    limit = 10000000
    # level = input("Sắp xếp theo: (1) Mới nhất trước, (2) Cũ nhất trước hoặc (3) Nhiều lượt xem trước? ")
    # if level == "1":
    #     level = "newest"
    # elif level == "2":
    #     level="oldest"
    # elif level == "3":
    #     level="popular"
    level="popular"

    t = time.localtime()
    timestamp = time.strftime('%Y-%m-%d_%H-%M-%S', t)

    resultFileName = (resultFolder + "/" + YTChannelID + "-" + timestamp + ".csv")
    resultExcelFileName = (resultFolder + "/" + YTChannelID + "-" + timestamp + ".xlsx")

    videos = scrapetube.get_channel(channel_url=searchString,limit=int(limit),sort_by=level)
    count = 0
    x = -1
    YTWatch_Prefix = 'https://www.youtube.com/watch?v='
    #print(videos)
    f = open(resultFileName, 'w', newline='', encoding="utf-8")
    writer = csv.writer(f)
    header = ['STT', 'Là video short?', 'Tên video','Lượt xem', 'sắp xếp Ngày đăng', 'Ngày đăng', '(sắp xếp) Thời lượng (giây)', 'Thời lượng','Link video']
    writer.writerow(header)

    for video in videos:
        title = video['title']['runs'][x+1]['text']
        isShort = "ko phải"
        #print(title)
        viewCount = video['viewCountText']['simpleText']
        viewCount = re.sub("\D","",viewCount)
        
        #print(viewCount)
        publishedTime = video['publishedTimeText']['simpleText'] 
        #print(publishedTime)
        videoLink = f'{YTWatch_Prefix}{video["videoId"]}'
        lengthText = video['lengthText']['simpleText']
        lengthSeconds = sum(int(x) * 60 ** i for i, x in enumerate(reversed(lengthText.split(':'))))
        count = count + 1
        line = [count,isShort,title,viewCount,count,publishedTime,lengthSeconds, lengthText, videoLink]
        writer.writerow(line)
        logText.insert(END, str(count) + " ---> " + videoLink + "\n", hyperlink.add(partial(webbrowser.open,videoLink)))
        logText.see(END)
    
    #
    PAGE_TOKEN = ""
    page = 1
    while 1:
        try:
            if PAGE_TOKEN == "":
                short_url = "https://yt0.lemnoslife.com/channels?part=shorts&id=" + YTChannelID
            else:
                short_url = "https://yt0.lemnoslife.com/channels?part=shorts&id=" + YTChannelID + "&pageToken=" + PAGE_TOKEN
            logText.insert(END, "\t---> Đang lấy danh sách video short (" +str(page) +"): "  + short_url + "\n\n", hyperlink.add(partial(webbrowser.open,short_url)))
            page = page + 1
            response = requests.get(short_url)
            PAGE_TOKEN = ""
            response.raise_for_status()
            if response.status_code == 200:
                jsonRes = response.json()
                #jsonRes = json.loads(response.text)
                #print(jsonRes)
                short_videos = jsonRes['items'][0]['shorts']
                for video in short_videos:
                    isShort = "đúng"
                    title = video['title']
                    #print(title)
                    viewCount = video['viewCount']
                    viewCount = re.sub("\D","", str(viewCount))
                    
                    #print(viewCount)
                    publishedTime = "không biết"
                    #print(publishedTime)
                    videoLink = f'{YTWatch_Prefix}{video["videoId"]}'
                    if video['duration'] == "1 minute, 1 second":
                        lengthSeconds = "61"
                        lengthText = "01:01"
                    elif video['duration'] == "1 minute":
                        lengthSeconds = "60"
                        lengthText = "01:00"
                    else:
                        lengthSeconds = re.sub(" seconds", "", str(video['duration']))
                        lengthText = "00:" + lengthSeconds
                    count = count + 1
                    line = [count,isShort,title,viewCount,count,publishedTime,lengthSeconds, lengthText, videoLink]
                    writer.writerow(line)
                    logText.insert(END, str(count) + " ---> " + videoLink + "\n", hyperlink.add(partial(webbrowser.open,videoLink)))
                    logText.see(END)
                if "nextPageToken" in response.text:
                    PAGE_TOKEN = jsonRes['items'][0]['nextPageToken']

        except Exception as err:
            print(f'Other error occurred: {err}')
            logText.insert(END, " ---> Có lỗi xảy ra: " + f'{err}' + "\n")
            logText.see(END)
            PAGE_TOKEN = ""
        
        if PAGE_TOKEN == "":
            break

    f.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    floats = [0, 3, 4, 6]

    with open(os.path.abspath(resultFileName), encoding='utf-8') as f1:
        reader = csv.reader(f1, delimiter=',')

        # fix error: writing csv to excel gives 'number formatted as text'
        # https://stackoverflow.com/a/45255614
        for row_index, row in enumerate(reader):
            for column_index, cell in enumerate(row):

                column_letter = get_column_letter((column_index + 1))

                if column_index in floats:
                    s = cell
                    #Handles heading row or non floats
                    try:
                        s = float(s)
                        ws[('%s%s'%(column_letter, (row_index + 1)))].value = s

                    except ValueError:
                        ws[('%s%s'%(column_letter, (row_index + 1)))].value = s

                elif column_index not in floats:
                    #Handles openpyxl 'illigal chars'
                    try:
                        ws[('%s%s'%(column_letter, (row_index + 1)))].value = cell
                    except:
                        ws[('%s%s'%(column_letter, (row_index + 1)))].value = 'illigal char'

            for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1):
                for cell in rows:
                    cell.font = Font(color="FFFFFF", bold=True)
                    cell.fill = PatternFill(bgColor="5badeb", fill_type = "mediumGray")
        f1.close()

    # Automatically adjust width of an excel file's columns
    # https://stackoverflow.com/a/39530676

    col_index = -1
    for col in ws.columns:
        col_index = col_index + 1
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 1)
        if adjusted_width > 50:
            adjusted_width = 50
        
        ws.column_dimensions[column].width = adjusted_width

    wb.save(os.path.abspath(resultExcelFileName))

    if os.path.exists(resultFileName):
        os.remove(resultFileName)
    else:
        print("The file does not exist")

    logText.insert(END, '\nXử lý xong ... kết quả được lưu tại: \n\t' + os.path.abspath(resultExcelFileName) + '\n')
    logText.see(END)
    #exploreFile(os.path.abspath(resultExcelFileName))
    os.startfile(os.path.abspath(resultExcelFileName))

def takeInput():
    try:
        inputText.delete("1.0","end-1c")
        logText.delete("1.0","end-1c")
    except Exception:
        print("clear text in InputTextBox failed")
    # inputText.insert("1.0", pyperclip.paste())
    # searchText = inputText.get("1.0", "end-1c")
    searchText = pyperclip.paste().strip()
    print(searchText)
   
    searchWithIDOnly = searchText.startswith(YTChannelID_Prefix, 0, 2)
    print(YTChannelID_Prefix + "------------>" + searchText)
    print(searchWithIDOnly)
    if searchWithIDOnly is True:
        searchText = "https://www.youtube.com/channel/" + searchText
        
    if YTChannelURL_Prefix in searchText:
        inputText.insert("1.0", searchText)
        logText.insert(END, '\t---> Đang xử lý kênh: '+ searchText + '\n', hyperlink.add(partial(webbrowser.open,searchText)))
        pyperclip.copy(searchText)
        getVideoList(searchText)
    else:
        inputText.insert("1.0", pyperclip.paste())
        logText.insert(END, "LỖI RỒI: ID kênh hoặc đường link không đúng định dạng.")
        logText.insert(END, "\nĐể lấy ID kênh:")
        logText.insert(END, "\n\t- Bước 1: BẤM VÀO ĐÂY để mở trang hỗ trợ https://seostudio.tools/youtube-channel-id", hyperlink.add(partial(webbrowser.open,"https://seostudio.tools/youtube-channel-id")))
        logText.insert(END, "\n\t- Bước 2: Copy đường link của kênh, cách làm: bấm vào tên kênh hoặc hình đại diện của kênh khi đang phát 1 video bất kỳ, link của kênh sẽ có chữ @ ở gần cuối dạng @xxxx), ví dụ: https://www.youtube.com/@tramcuuho_traitim")
        logText.insert(END, "\n\t- Bước 3: Dán đường link của kênh vào, và bấm nút 'Find Now'")
        logText.insert(END, "\n\t- Bước 4: Copy ID kênh ở dòng Channel ID, bắt đầu bằng chữ UC, ví dụ: UCsSrvUHWzNt7yXWxzOkyjhw")

titleLabel = Label(text="Nhập ID của 1 kênh Youtube (bắt đầu bằng chữ UC, ví dụ: UCsSrvUHWzNt7yXWxzOkyjhw)\nhoặc đường link kênh (có dạng https://www.youtube.com/channel/ID_kênh)", font= ('Arial', 16, 'bold'))

inputText = Text(root, height=5,
                width=150,
                font= ('Arial', 14),
                bg="light yellow")

logText = ScrolledText(root, height=7,
              width= 150,
              font= ('Arial', 13),
              bg="light cyan")

searchBtn = Button(root, height=2,
                 width=30,
                 text="Dán và lấy danh sách video",
                 font= ('Arial', 14, 'bold'),
                 command=lambda: takeInput())

openResultFolderBtn = Button(root, height=2,
                 width=30,
                 text="Mở thư mục kết quả",
                 font= ('Arial', 14, 'bold'),
                 command=lambda: explore())

getAvatarBannerBtn = Button(root, height=2,
                 width=30,
                 text="Tải hình avatar và banner của kênh",
                 font= ('Arial', 14, 'bold'),
                 command=lambda: callback("https://imageyoutube.com/?"))

titleLabel.pack()
inputText.pack()
searchBtn.pack()
openResultFolderBtn.pack()
getAvatarBannerBtn.pack()
logText.pack()

hyperlink= HyperlinkManager(logText)

mainloop()
