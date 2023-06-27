from tkinter import *
import webbrowser
from functools import partial

import os
import scrapetube
import pyperclip
import csv
import time
import re
import threading

import openpyxl
from openpyxl.utils import get_column_letter

from tkinter.scrolledtext import ScrolledText

# from http://effbot.org/zone/tkinter-text-hyperlink.htm
class HyperlinkManager:

    def __init__(self, text):

        self.text = text

        self.text.tag_config("hyper", foreground="blue", underline=1)

        self.text.tag_bind("hyper", "<Enter>", self._enter)
        self.text.tag_bind("hyper", "<Leave>", self._leave)
        self.text.tag_bind("hyper", "<Button-1>", self._click)

        self.reset()

    def reset(self):
        self.links = {}

    def add(self, action):
        # add an action to the manager.  returns tags to use in
        # associated text widget
        tag = "hyper-%d" % len(self.links)
        self.links[tag] = action
        return "hyper", tag

    def _enter(self, event):
        self.text.config(cursor="hand2")

    def _leave(self, event):
        self.text.config(cursor="")

    def _click(self, event):
        for tag in self.text.tag_names(CURRENT):
            if tag[:6] == "hyper-":
                self.links[tag]()
                return

import subprocess
FILEBROWSER_PATH = os.path.join(os.getenv('WINDIR'), 'explorer.exe')

# Define a callback function
def callback(url):
   webbrowser.open_new_tab(url)

root = Tk()
root.geometry("1000x500")
root.title("Lấy danh sách video của 1 kênh Youtube - Tải hình avatar và banner")

YTChannelURL_Prefix = "https://www.youtube.com/channel/"
YTChannelID_Prefix = "UC"
resultFolder = "KetQua"

try:
    os.mkdir(resultFolder)
except FileExistsError:
    print("Folder already exists")

def explore():
    path = os.path.abspath(resultFolder)
    # explorer would choke on forward slashes
    path = os.path.normpath(path)

    if os.path.isdir(path):
        subprocess.run([FILEBROWSER_PATH, path])
    elif os.path.isfile(path):
        subprocess.run([FILEBROWSER_PATH, '/select,', path])

def exploreFile(path):
    if os.path.isdir(path):
        subprocess.run([FILEBROWSER_PATH, path])
    elif os.path.isfile(path):
        subprocess.run([FILEBROWSER_PATH, '/select,', path])        

# def getAvatarBanner():
#     webbrowser.open("https://imageyoutube.com/?")

def getVideoList(searchString):
    x = threading.Thread(target=getVideoListPrivate, args=(searchString,))
    x.start()
    # x.join()

def getVideoListPrivate(searchString):
    parts = searchString.split(YTChannelURL_Prefix, 1)
    #Output.insert(END, 'ID kênh: '+ INPUT_parts[1] +'\n')
    YTChannelID = parts[1]
    
    #channel_url = input("Nhap duong link kenh Youtube (dinh dang https://www.youtube.com/channel/xxxxx), de lay xxxxx DUNG TRANG https://http5.org/chan: ")
    # limit = input("Muon lay bao nhieu video? ")
    limit = 10000000
    # level = input("Sắp xếp theo: (1) Mới nhất trước, (2) Cũ nhất trước hoặc (3) Nhiều lượt xem trước? ")
    # if level == "1":
    #     level = "newest"
    # elif level == "2":
    #     level="oldest"
    # elif level == "3":
    #     level="popular"
    level="popular"

    t = time.localtime()
    timestamp = time.strftime('%Y-%m-%d_%H-%M-%S', t)

    resultFileName = (resultFolder + "/" + YTChannelID + "-" + timestamp + ".csv")
    resultExcelFileName = (resultFolder + "/" + YTChannelID + "-" + timestamp + ".xlsx")

    videos = scrapetube.get_channel(channel_url=searchString,limit=int(limit),sort_by=level)
    count = 0
    x = -1
    YTWatch_Prefix = 'https://www.youtube.com/watch?v='
    #print(videos)
    f = open(resultFileName, 'w', newline='', encoding="utf-8")
    writer = csv.writer(f)
    header = ['STT', 'Tên video','Lượt xem', 'sắp xếp Ngày đăng', 'Ngày đăng', 'Link video']
    writer.writerow(header)

    for video in videos:
        title = video['title']['runs'][x+1]['text']
        #print(title)
        viewCount = video['viewCountText']['simpleText']
        viewCount = re.sub("\D","",viewCount)
        
        #print(viewCount)
        publishedTime = video['publishedTimeText']['simpleText'] 
        #print(publishedTime)
        videoLink = f'{YTWatch_Prefix}{video["videoId"]}'
        count = count + 1
        line = [count,title,viewCount,count,publishedTime,videoLink]
        writer.writerow(line)
        logText.insert(END, str(count) + " ---> " + videoLink + "\n", hyperlink.add(partial(webbrowser.open,videoLink)))
        logText.see(END)

    f.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    floats = [0, 2, 3]

    with open(os.path.abspath(resultFileName), encoding='utf-8') as f1:
        reader = csv.reader(f1, delimiter=',')

        # fix error: writing csv to excel gives 'number formatted as text'
        # https://stackoverflow.com/a/45255614
        for row_index, row in enumerate(reader):
            for column_index, cell in enumerate(row):

                column_letter = get_column_letter((column_index + 1))

                if column_index in floats:
                    s = cell
                    #Handles heading row or non floats
                    try:
                        s = float(s)
                        ws[('%s%s'%(column_letter, (row_index + 1)))].value = s

                    except ValueError:
                        ws[('%s%s'%(column_letter, (row_index + 1)))].value = s

                elif column_index not in floats:
                    #Handles openpyxl 'illigal chars'
                    try:
                        ws[('%s%s'%(column_letter, (row_index + 1)))].value = cell
                    except:
                        ws[('%s%s'%(column_letter, (row_index + 1)))].value = 'illigal char'
        f1.close()

    # Automatically adjust width of an excel file's columns
    # https://stackoverflow.com/a/39530676
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        if adjusted_width > 50:
            adjusted_width = 50
        ws.column_dimensions[column].width = adjusted_width

    wb.save(os.path.abspath(resultExcelFileName))

    if os.path.exists(resultFileName):
        os.remove(resultFileName)
    else:
        print("The file does not exist")

    logText.insert(END, '\nXử lý xong ... kết quả được lưu tại: \n\t' + os.path.abspath(resultExcelFileName) + '\n')
    logText.see(END)
    #exploreFile(os.path.abspath(resultExcelFileName))
    os.startfile(os.path.abspath(resultExcelFileName))

def takeInput():
    try:
        inputText.delete("1.0","end-1c")
        logText.delete("1.0","end-1c")
    except Exception:
        print("clear text in InputTextBox failed")
    # inputText.insert("1.0", pyperclip.paste())
    # searchText = inputText.get("1.0", "end-1c")
    searchText = pyperclip.paste().strip()
    print(searchText)
   
    searchWithIDOnly = searchText.startswith(YTChannelID_Prefix, 0, 2)
    print(YTChannelID_Prefix + "------------>" + searchText)
    print(searchWithIDOnly)
    if searchWithIDOnly is True:
        searchText = "https://www.youtube.com/channel/" + searchText
        
    if YTChannelURL_Prefix in searchText:
        inputText.insert("1.0", searchText)
        logText.insert(END, '\t---> Đang xử lý kênh: '+ searchText + '\n', hyperlink.add(partial(webbrowser.open,searchText)))
        pyperclip.copy(searchText)
        getVideoList(searchText)
    else:
        inputText.insert("1.0", pyperclip.paste())
        logText.insert(END, "LỖI RỒI: ID kênh hoặc đường link không đúng định dạng.")
        logText.insert(END, "\nĐể lấy ID kênh:")
        logText.insert(END, "\n\t- Bước 1: BẤM VÀO ĐÂY để mở trang hỗ trợ https://http5.org/chan", hyperlink.add(partial(webbrowser.open,"https://http5.org/chan")))
        logText.insert(END, "\n\t- Bước 2: dán đường link 1 video bất kỳ của kênh vào, và bấm nút 'Phân tách Kênh hoặc Parse Channel'")
        logText.insert(END, "\n\t- Bước 3: Copy ID kênh ở dòng ID, bắt đầu bằng chữ UC, ví dụ: UCsSrvUHWzNt7yXWxzOkyjhw ")

titleLabel = Label(text="Nhập ID của 1 kênh Youtube (bắt đầu bằng chữ UC, ví dụ: UCsSrvUHWzNt7yXWxzOkyjhw)\nhoặc đường link kênh (có dạng https://www.youtube.com/channel/ID_kênh)", font= ('Arial', 16, 'bold'))

inputText = Text(root, height=5,
                width=150,
                font= ('Arial', 14),
                bg="light yellow")

logText = ScrolledText(root, height=7,
              width= 150,
              font= ('Arial', 13),
              bg="light cyan")

searchBtn = Button(root, height=2,
                 width=30,
                 text="Dán và lấy danh sách video",
                 font= ('Arial', 14, 'bold'),
                 command=lambda: takeInput())

openResultFolderBtn = Button(root, height=2,
                 width=30,
                 text="Mở thư mục kết quả",
                 font= ('Arial', 14, 'bold'),
                 command=lambda: explore())

getAvatarBannerBtn = Button(root, height=2,
                 width=30,
                 text="Tải hình avatar và banner của kênh",
                 font= ('Arial', 14, 'bold'),
                 command=lambda: callback("https://imageyoutube.com/?"))

titleLabel.pack()
inputText.pack()
searchBtn.pack()
openResultFolderBtn.pack()
getAvatarBannerBtn.pack()
logText.pack()

hyperlink= HyperlinkManager(logText)

mainloop()
